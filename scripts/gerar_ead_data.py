#!/usr/bin/env python3
"""Gera assets/ead-data.js a partir dos relatorios do SGN/FIESC e das grades curriculares.

Uso: python gerar_ead_data.py [pasta_com_relatorios]
A pasta padrao e ../../Claude_gestao/data/incoming (relativa a este arquivo).
"""
import csv
import datetime
import glob
import io
import json
import os
import re
import sys
import unicodedata
import urllib.request
from collections import defaultdict

import openpyxl

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SCRIPTS_DIR = os.path.dirname(os.path.abspath(__file__))
NAME_MAP = {"Técnico em Internet das Coisas - IoT": "Técnico em Internet das Coisas"}


def norm(s):
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFD", s)
    return "".join(c for c in s if unicodedata.category(c) != "Mn")


def parse_num(v):
    if v in (None, ""):
        return None
    try:
        return float(str(v).replace(",", "."))
    except ValueError:
        return None


def fetch_csv(spreadsheet_id, gid):
    url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=csv&gid={gid}"
    with urllib.request.urlopen(url, timeout=30) as resp:
        data = resp.read().decode("utf-8")
    return list(csv.reader(io.StringIO(data)))


def parse_grade(rows):
    course_name = None
    for r in rows[:3]:
        for cell in r:
            if cell and "Técnico" in cell:
                course_name = cell.split("\n")[0].strip()
                break
        if course_name:
            break
    header_idx = next(i for i, r in enumerate(rows) if any(c.strip() == "Unidades curriculares" for c in r))
    header = rows[header_idx]
    col = {c.strip(): i for i, c in enumerate(header) if c.strip()}
    uc_col = col["Unidades curriculares"] + 1
    total_col = col["Carga Horária Total"]
    pres_col = col["Carga Horária Presencial"]
    ead_ucs = {}
    for r in rows[header_idx + 1:]:
        if uc_col >= len(r):
            continue
        uc = r[uc_col].strip()
        total = parse_num(r[total_col]) if total_col < len(r) else None
        pres = parse_num(r[pres_col]) if pres_col < len(r) else None
        if not uc or total is None or pres is None:
            continue
        if uc.lower().startswith("versão") or "carga horária" in uc.lower():
            continue
        if pres == 0 and total > 0:
            ead_ucs[norm(uc)] = {"uc": uc, "cargaHoraria": int(total)}
    return course_name, ead_ucs


def find_latest(prefix, directory):
    candidates = glob.glob(os.path.join(directory, prefix + "*.xlsx"))

    def extract_date(path):
        m = re.search(r"(\d{2})(\d{2})(\d{4})", os.path.basename(path))
        if not m:
            return datetime.date.min
        d, mth, y = m.groups()
        return datetime.date(int(y), int(mth), int(d))

    if not candidates:
        return None
    return max(candidates, key=extract_date)


def parse_set(v):
    if not v:
        return []
    v = v.strip()
    if v.startswith("{") and v.endswith("}"):
        v = v[1:-1]
    if not v.strip():
        return []
    items = re.findall(r'"([^"]*)"|([^,]+)', v)
    return [(a or b).strip() for a, b in items if (a or b).strip()]


def escolher_diario(diarios, hoje):
    def status(d):
        if d["inicio"] is None or d["fim"] is None:
            return "sem-data"
        ini_d, fim_d = d["inicio"].date(), d["fim"].date()
        if fim_d < hoje:
            return "concluida"
        if ini_d > hoje:
            return "futura"
        return "andamento"

    andamento = [d for d in diarios if status(d) == "andamento"]
    if andamento:
        return andamento[0]
    futuras = [d for d in diarios if status(d) == "futura"]
    if futuras:
        return min(futuras, key=lambda d: d["inicio"])
    concluidas = [d for d in diarios if status(d) == "concluida"]
    if concluidas:
        return max(concluidas, key=lambda d: d["fim"])
    return None


def main():
    default_incoming = os.path.join(BASE_DIR, "..", "Claude_gestao", "data", "incoming")
    incoming_dir = os.path.abspath(sys.argv[1] if len(sys.argv) > 1 else default_incoming)

    with open(os.path.join(SCRIPTS_DIR, "etg_turmas.json"), encoding="utf-8") as f:
        target_ids = set(json.load(f)["ids"])

    with open(os.path.join(SCRIPTS_DIR, "grade_sources.json"), encoding="utf-8") as f:
        grade_sources = json.load(f)

    print(f"Baixando {len(grade_sources['cursos'])} grades curriculares...")
    course_ead = {}
    for item in grade_sources["cursos"]:
        rows = fetch_csv(item["spreadsheetId"], item["gid"])
        _, ead_ucs = parse_grade(rows)
        course_ead[item["curso"]] = ead_ucs
        print(f"  {item['curso']}: {len(ead_ucs)} UCs 100% EAD")

    turmas_path = find_latest("relatorio_turmas_em_andamento_", incoming_dir)
    diarios_path = find_latest("relatorio_diarios_em_andamento_", incoming_dir)
    matricula_path = find_latest("relatorio_matricula_situacao_UCs_", incoming_dir)
    if not (turmas_path and diarios_path and matricula_path):
        raise SystemExit(f"Relatorios nao encontrados em {incoming_dir}")
    print(f"Turmas: {os.path.basename(turmas_path)}")
    print(f"Diarios: {os.path.basename(diarios_path)}")
    print(f"Matricula: {os.path.basename(matricula_path)}")

    d, mth, y = re.search(r"(\d{2})(\d{2})(\d{4})", os.path.basename(turmas_path)).groups()
    gerado_em = f"{y}-{mth}-{d}"

    wb = openpyxl.load_workbook(turmas_path, read_only=True, data_only=True)
    ws = wb["Dados"] if "Dados" in wb.sheetnames else wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    idx = {n: i for i, n in enumerate(header)}
    turmas = {}
    for row in rows_iter:
        idt = row[idx["id_turma"]]
        if idt is None or int(idt) not in target_ids:
            continue
        idt = int(idt)
        produto = row[idx["nome_produto"]]
        curso = NAME_MAP.get(produto, produto)
        if curso not in course_ead:
            print(f"  AVISO: curso '{curso}' (turma {idt}) sem grade cadastrada, pulando")
            continue
        turmas[idt] = {
            "id": idt,
            "nome": row[idx["nome_turma"]],
            "curso": curso,
            "unidade": (row[idx["unidade_execucao"]] or "").replace("SENAI/SC - ", ""),
            "link": f"https://sgn.sesisenai.org.br/pages/execucaoEducacao/execucao-educacao.html?idTurma={idt}",
        }
    wb.close()
    faltando = target_ids - set(turmas.keys())
    if faltando:
        print(f"  AVISO: turmas ETG configuradas mas nao encontradas no relatorio: {sorted(faltando)}")

    wb = openpyxl.load_workbook(diarios_path, read_only=True, data_only=True)
    ws = wb["Planilha1"] if "Planilha1" in wb.sheetnames else wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    idx = {n: i for i, n in enumerate(header)}
    diarios_by_turma = defaultdict(lambda: defaultdict(list))
    for row in rows_iter:
        idt = row[idx["id_turma"]]
        if idt is None or int(idt) not in turmas:
            continue
        if norm(row[idx["situacao_diario"]]) != "em andamento":
            continue
        idt = int(idt)
        uc_original = row[idx["unidade_curricular"]]
        uc_norm = norm(uc_original)
        diarios_by_turma[idt][uc_norm].append({
            "uc": uc_original,
            "inicio": row[idx["inicio_diario"]],
            "fim": row[idx["termino_diario"]],
        })
    wb.close()

    wb = openpyxl.load_workbook(matricula_path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    idx = {n: i for i, n in enumerate(header)}
    matricula_by_turma = defaultdict(list)
    for row in rows_iter:
        idt = row[idx["id_turma"]]
        if idt is None or int(idt) not in turmas:
            continue
        if row[idx["situacao_matricula"]] != "Matriculado / Regular":
            continue
        idt = int(idt)
        matricula_by_turma[idt].append({
            "cursando": {norm(x) for x in parse_set(row[idx["uc_cursando"]])},
            "aprovadas": {norm(x) for x in parse_set(row[idx["uc_aprovadas"]])},
            "reprovada": {norm(x) for x in parse_set(row[idx["uc_reprovada"]])},
        })
    wb.close()

    hoje = datetime.date.today()
    output_turmas = []
    for idt, t in sorted(turmas.items(), key=lambda kv: (kv[1]["curso"], kv[1]["nome"])):
        ead_ucs = course_ead[t["curso"]]
        regs = matricula_by_turma.get(idt, [])
        ucs_out = []
        for uc_norm, info in ead_ucs.items():
            diarios = diarios_by_turma.get(idt, {}).get(uc_norm, [])
            escolhido = escolher_diario(diarios, hoje)
            inicio = escolhido["inicio"].date().isoformat() if escolhido and escolhido["inicio"] else None
            fim = escolhido["fim"].date().isoformat() if escolhido and escolhido["fim"] else None
            ucs_out.append({
                "uc": info["uc"],
                "cargaHoraria": info["cargaHoraria"],
                "inicio": inicio,
                "fim": fim,
                "regulares": len(regs),
                "cursando": sum(1 for s in regs if uc_norm in s["cursando"]),
                "aprovados": sum(1 for s in regs if uc_norm in s["aprovadas"]),
                "reprovados": sum(1 for s in regs if uc_norm in s["reprovada"]),
            })
        ucs_out.sort(key=lambda u: (u["inicio"] or "9999", u["uc"]))

        diario_turma_out = []
        for uc_norm, diarios in diarios_by_turma.get(idt, {}).items():
            escolhido = escolher_diario(diarios, hoje)
            if not escolhido:
                continue
            diario_turma_out.append({
                "uc": escolhido["uc"],
                "inicio": escolhido["inicio"].date().isoformat() if escolhido["inicio"] else None,
                "fim": escolhido["fim"].date().isoformat() if escolhido["fim"] else None,
            })

        output_turmas.append({
            "id": t["id"], "nome": t["nome"], "curso": t["curso"],
            "unidade": t["unidade"], "link": t["link"], "ucs": ucs_out,
            "diarioTurma": diario_turma_out,
        })

    output = {"geradoEm": gerado_em, "turmas": output_turmas}
    out_path = os.path.join(BASE_DIR, "assets", "ead-data.js")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("window.EAD_OFERTAS = " + json.dumps(output, ensure_ascii=False, indent=2) + ";\n")
    print(f"Gravado {out_path} ({len(output_turmas)} turmas)")


if __name__ == "__main__":
    main()
