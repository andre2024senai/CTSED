#!/usr/bin/env python3
"""Gera o JSON privado (com nomes) de alunos por turma ETG, para publicar no
Realtime Database do projeto Firebase ctsed-senai-etg (nao no site publico).

Uso: python gerar_alunos_privado.py [pasta_com_relatorios] [arquivo_saida.json]
"""
import datetime
import glob
import json
import os
import re
import sys
import unicodedata
from collections import defaultdict

import openpyxl

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SCRIPTS_DIR = os.path.dirname(os.path.abspath(__file__))
NAME_MAP = {"Técnico em Internet das Coisas - IoT": "Técnico em Internet das Coisas"}


def norm(s):
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFD", s)
    return "".join(c for c in s if unicodedata.category(c) != "Mn")


def find_latest(prefix, directory):
    candidates = glob.glob(os.path.join(directory, prefix + "*.xlsx"))

    def extract_date(path):
        m = re.search(r"(\d{2})(\d{2})(\d{4})", os.path.basename(path))
        if not m:
            return datetime.date.min
        d, mth, y = m.groups()
        return datetime.date(int(y), int(mth), int(d))

    if not candidates:
        return None
    return max(candidates, key=extract_date)


def parse_set(v):
    if not v:
        return []
    v = v.strip()
    if v.startswith("{") and v.endswith("}"):
        v = v[1:-1]
    if not v.strip():
        return []
    items = re.findall(r'"([^"]*)"|([^,]+)', v)
    return [(a or b).strip() for a, b in items if (a or b).strip()]


def parse_phones(value):
    candidates = re.findall(r"\d{10,13}", str(value or ""))
    output = []
    for candidate in candidates:
        digits = candidate
        if len(digits) in (10, 11):
            digits = "55" + digits
        if len(digits) in (12, 13) and digits not in output:
            output.append(digits)
    return output


def main():
    default_incoming = os.path.join(BASE_DIR, "..", "Claude_gestao", "data", "incoming")
    incoming_dir = os.path.abspath(sys.argv[1] if len(sys.argv) > 1 else default_incoming)
    out_path = sys.argv[2] if len(sys.argv) > 2 else os.path.join(BASE_DIR, "scripts", "alunos-privado.json")

    with open(os.path.join(SCRIPTS_DIR, "etg_turmas.json"), encoding="utf-8") as f:
        target_ids = set(json.load(f)["ids"])

    turmas_path = find_latest("relatorio_turmas_em_andamento_", incoming_dir)
    matricula_path = find_latest("relatorio_matricula_situacao_UCs_", incoming_dir)
    if not (turmas_path and matricula_path):
        raise SystemExit(f"Relatorios nao encontrados em {incoming_dir}")
    print(f"Turmas: {os.path.basename(turmas_path)}")
    print(f"Matricula: {os.path.basename(matricula_path)}")

    d, mth, y = re.search(r"(\d{2})(\d{2})(\d{4})", os.path.basename(turmas_path)).groups()
    gerado_em = f"{y}-{mth}-{d}"

    wb = openpyxl.load_workbook(turmas_path, read_only=True, data_only=True)
    ws = wb["Dados"] if "Dados" in wb.sheetnames else wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    idx = {n: i for i, n in enumerate(header)}
    turmas = {}
    for row in rows_iter:
        idt = row[idx["id_turma"]]
        if idt is None or int(idt) not in target_ids:
            continue
        idt = int(idt)
        produto = row[idx["nome_produto"]]
        turmas[idt] = {
            "id": idt,
            "nome": row[idx["nome_turma"]],
            "curso": NAME_MAP.get(produto, produto),
            "unidade": (row[idx["unidade_execucao"]] or "").replace("SENAI/SC - ", ""),
            "link": f"https://sgn.sesisenai.org.br/pages/execucaoEducacao/execucao-educacao.html?idTurma={idt}",
            "alunos": [],
        }
    wb.close()

    wb = openpyxl.load_workbook(matricula_path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    idx = {n: i for i, n in enumerate(header)}
    count_alunos = 0
    for row in rows_iter:
        idt = row[idx["id_turma"]]
        if idt is None or int(idt) not in turmas:
            continue
        idt = int(idt)
        situacao = str(row[idx["situacao_matricula"]] or "")
        if norm(situacao) != "matriculado / regular":
            continue
        turmas[idt]["alunos"].append({
            "nome": str(row[idx["aluno"]] or "").strip(),
            "matricula": str(row[idx["id_matricula"]] or ""),
            "status": situacao,
            "email": str(row[idx["email_contato"]] or "").strip(),
            "telefones": parse_phones(row[idx["telefones"]]),
            "ucsReprovadas": parse_set(row[idx["uc_reprovada"]]),
            "ucsCursando": parse_set(row[idx["uc_cursando"]]),
            "ucsAprovadas": parse_set(row[idx["uc_aprovadas"]]),
        })
        count_alunos += 1
    wb.close()

    for turma in turmas.values():
        turma["alunos"].sort(key=lambda a: a["nome"])

    output = {"geradoEm": gerado_em, "turmas": {str(k): v for k, v in turmas.items()}}
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    print(f"Gravado {out_path} ({len(turmas)} turmas, {count_alunos} alunos regulares)")


if __name__ == "__main__":
    main()
